import requests
from bs4 import BeautifulSoup
import openpyxl

# URL ---> https://www.imdb.com/chart/top

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank','Movie Name','Year of Release','IMDB Rating'])


try:
   source = requests.get('https://www.imdb.com/chart/top')
   source.raise_for_status()

   soup = BeautifulSoup(source.text,'html.parser')
   #print(soup.prettify())
   
   movies = soup.find('tbody',class_='lister-list').find_all('tr')
 #  print(len(movies))

   for mov in movies:
     name = mov.find('td',class_='titleColumn').a.text
     rank = mov.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]
     year = mov.find('td',class_='titleColumn').span.text.strip()
     rating = mov.find('td',class_="ratingColumn imdbRating").strong.text
     print(rank,name,year,rating)
     sheet.append([rank,name,year,rating])

except Exception as e:
    print(e)
  
excel.save('IMBD Movie Rating.xlsx')